from flask import Flask, render_template, request, redirect, flash, session, make_response, abort
from supabase import create_client, Client
from dotenv import load_dotenv
from datetime import datetime, timezone, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

import csv
import io
import os
import secrets
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

load_dotenv()

app = Flask(__name__)


def _env_to_bool(value, default=False):
    if value is None:
        return default
    return str(value).strip().lower() in {'1', 'true', 't', 'yes', 'y', 'on'}


secret_key = os.environ.get('SECRET_KEY')
if not secret_key:
    raise RuntimeError('SECRET_KEY is obrigatoria. Defina a variavel de ambiente SECRET_KEY.')

app.config['SECRET_KEY'] = secret_key
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = _env_to_bool(os.environ.get('SESSION_COOKIE_SECURE'))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

PRIORIDADES_VALIDAS = ['Alta', 'Média', 'Baixa']
ORDEM_PRIORIDADE = {'Alta': 1, 'Média': 2, 'Baixa': 3}
SLA_DIAS_POR_PRIORIDADE = {'Alta': 2, 'Média': 5, 'Baixa': 10}
STATUS_DEMANDA_VALIDOS = ['Aberta', 'Em andamento', 'Parada', 'Finalizada']
STATUS_TRANSITIONS = {
    'Aberta': {'Em andamento', 'Parada'},
    'Em andamento': {'Parada', 'Finalizada'},
    'Parada': {'Em andamento', 'Finalizada'},
    'Finalizada': {'Aberta'},
}
DIAS_PARADA = 3  # configurável
ORDENACOES_LISTAGEM = {
    'updated_at': 'updated_at',
    'prioridade': 'prioridade',
    'due_date': 'due_date',
}
ORDENS_DIRECAO = {'asc', 'desc'}
ITENS_POR_PAGINA_PADRAO = 12
ITENS_POR_PAGINA_MAX = 50


# ---------------------------------------------------------------------------
# Helpers de domínio
# ---------------------------------------------------------------------------

def _parse_iso_datetime(data_str):
    if not data_str:
        return None

    try:
        return datetime.fromisoformat(data_str.replace('Z', '+00:00'))
    except Exception:
        return None


def _formatar_data_hora(data_str):
    data = _parse_iso_datetime(data_str)
    return data.strftime('%d/%m/%Y %H:%M') if data else '-'


def calcular_due_date(prioridade, data_base=None):
    data_referencia = data_base or datetime.now(timezone.utc)
    dias_sla = SLA_DIAS_POR_PRIORIDADE.get(prioridade, SLA_DIAS_POR_PRIORIDADE['Média'])
    return data_referencia + timedelta(days=dias_sla)


def calcular_dias_parada(demanda):
    """Retorna dias desde que a demanda entrou em Parada."""
    status = demanda.get('status') or 'Aberta'
    if status != 'Parada':
        return 0

    data_base = demanda.get('status_updated_at') or demanda.get('updated_at') or demanda.get('data_criacao')
    data = _parse_iso_datetime(data_base)
    if not data:
        return 0

    agora = datetime.now(timezone.utc)
    return max((agora - data).days, 0)


def calcular_indicadores_prazo(demanda):
    status = demanda.get('status') or 'Aberta'
    due_dt = _parse_iso_datetime(demanda.get('due_date'))
    agora = datetime.now(timezone.utc)

    dias_para_vencimento = None
    dias_atraso = 0
    esta_atrasada = False

    if due_dt:
        dias_para_vencimento = (due_dt - agora).days
        if status != 'Finalizada' and due_dt < agora:
            esta_atrasada = True
            dias_atraso = max((agora - due_dt).days, 0)

    demanda['due_date_fmt'] = _formatar_data_hora(demanda.get('due_date'))
    demanda['status_updated_at_fmt'] = _formatar_data_hora(demanda.get('status_updated_at'))
    demanda['resolved_at_fmt'] = _formatar_data_hora(demanda.get('resolved_at'))
    demanda['dias_para_vencimento'] = dias_para_vencimento
    demanda['dias_atraso'] = dias_atraso
    demanda['esta_atrasada'] = esta_atrasada


def preparar_demandas(demandas):
    dados = sorted(demandas or [], key=lambda d: (
        ORDEM_PRIORIDADE.get(d.get('prioridade', 'Média'), 2),
        d.get('data_criacao', '')
    ))
    for demanda in dados:
        demanda['dias_parada'] = calcular_dias_parada(demanda)
        calcular_indicadores_prazo(demanda)
        assignee = demanda.get('assignee')
        assignee_nome = None
        if isinstance(assignee, dict):
            assignee_nome = assignee.get('nome')
        if not assignee_nome and demanda.get('assignee_nome'):
            assignee_nome = demanda.get('assignee_nome')
        if not assignee_nome and demanda.get('assignee_id') == demanda.get('usuario_id'):
            assignee_nome = demanda.get('solicitante')
        demanda['assignee_nome'] = assignee_nome or 'Nao atribuido'
    return dados


def listar_usuarios():
    resposta = (
        supabase
        .table('usuarios')
        .select('id,nome')
        .order('nome')
        .execute()
    )
    return resposta.data or []


def listar_solicitantes(demandas):
    return sorted({
        demanda.get('solicitante')
        for demanda in demandas or []
        if demanda.get('solicitante')
    }, key=str.lower)


def parse_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalizar_status_filtros(status_values):
    return [status for status in status_values if status in STATUS_DEMANDA_VALIDOS]


def _parse_filtros_listagem(args):
    status_filtro = _normalizar_status_filtros(args.getlist('status'))
    if not status_filtro:
        status_unico = args.get('status', '').strip()
        if status_unico in STATUS_DEMANDA_VALIDOS:
            status_filtro = [status_unico]

    page = max(parse_int(args.get('page', 1), 1), 1)
    per_page_raw = max(parse_int(args.get('per_page', ITENS_POR_PAGINA_PADRAO), ITENS_POR_PAGINA_PADRAO), 1)
    per_page = min(per_page_raw, ITENS_POR_PAGINA_MAX)

    sort_by = args.get('sort_by', 'updated_at')
    if sort_by not in ORDENACOES_LISTAGEM:
        sort_by = 'updated_at'

    sort_dir = args.get('sort_dir', 'desc').lower()
    if sort_dir not in ORDENS_DIRECAO:
        sort_dir = 'desc'

    return {
        'filtro_prioridade': args.get('prioridade', 'Todas'),
        'filtro_solicitante': args.get('solicitante', '').strip(),
        'status_filtro': status_filtro,
        'periodo_inicio': args.get('data_inicio', '').strip(),
        'periodo_fim': args.get('data_fim', '').strip(),
        'assignee_id': args.get('assignee_id', '').strip(),
        'minhas_demandas': args.get('minhas_demandas') == '1',
        'sort_by': sort_by,
        'sort_dir': sort_dir,
        'page': page,
        'per_page': per_page,
    }


def gerar_relatorio_solicitantes(demandas):
    resumo = {}
    for demanda in demandas or []:
        chave = demanda.get('solicitante') or 'Sem solicitante'
        item = resumo.setdefault(chave, {
            'solicitante': chave,
            'total': 0,
            'alta': 0,
            'media': 0,
            'baixa': 0,
            'paradas': 0,
            'atrasadas': 0,
        })
        prioridade = demanda.get('prioridade')
        item['total'] += 1
        if prioridade == 'Alta':
            item['alta'] += 1
        elif prioridade == 'Média':
            item['media'] += 1
        elif prioridade == 'Baixa':
            item['baixa'] += 1
        demanda_preparada = dict(demanda)
        demanda_preparada['dias_parada'] = calcular_dias_parada(demanda_preparada)
        calcular_indicadores_prazo(demanda_preparada)
        if demanda_preparada.get('dias_parada', 0) >= DIAS_PARADA:
            item['paradas'] += 1
        if demanda_preparada.get('esta_atrasada'):
            item['atrasadas'] += 1

    return sorted(resumo.values(), key=lambda item: (-item['total'], item['solicitante'].lower()))


def usuario_pode_gerenciar(demanda):
    if not demanda:
        return False

    return (
        demanda.get('usuario_id') == session.get('usuario_id')
        or session.get('role') == 'manager'
    )


def usuario_pode_alterar_status(demanda):
    if not demanda:
        return False

    return (
        demanda.get('usuario_id') == session.get('usuario_id')
        or demanda.get('assignee_id') == session.get('usuario_id')
        or session.get('role') == 'manager'
    )


def transicao_status_valida(status_atual, novo_status):
    if status_atual == novo_status:
        return True

    return novo_status in STATUS_TRANSITIONS.get(status_atual, set())


def filtrar_criticas(demandas):
    """Retorna demandas de prioridade Alta E com dias_parada >= DIAS_PARADA."""
    return [
        d for d in (demandas or [])
        if d.get('prioridade') == 'Alta'
        and d.get('dias_parada', 0) >= DIAS_PARADA
    ]


def buscar_demanda(id):
    resposta = supabase.table('demandas').select('*').eq('id', id).execute()
    return resposta.data[0] if resposta.data else None


def registrar_evento_demanda(demanda_id, tipo, before_data=None, after_data=None, autor_id=None):
    if not demanda_id:
        return

    evento = {
        'demanda_id': demanda_id,
        'tipo': tipo,
        'autor_id': autor_id if autor_id is not None else session.get('usuario_id'),
        'before_data': before_data or {},
        'after_data': after_data or {},
    }
    supabase.table('demanda_eventos').insert(evento).execute()


def listar_eventos_demanda(demanda_id):
    resposta = (
        supabase
        .table('demanda_eventos')
        .select('id,tipo,before_data,after_data,created_at,autor_id,autor:autor_id(nome)')
        .eq('demanda_id', demanda_id)
        .order('created_at', desc=True)
        .execute()
    )
    return resposta.data or []


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not session.get('usuario_id'):
            flash('Faça login para continuar')
            return redirect('/login')
        return view_func(*args, **kwargs)
    return wrapped_view


def manager_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if session.get('role') != 'manager':
            flash('Acesso permitido apenas para perfil gerencial.')
            return redirect('/dashboard')
        return view_func(*args, **kwargs)
    return wrapped_view


def _get_or_create_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


@app.before_request
def csrf_protect():
    if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        session_token = session.get('_csrf_token')
        request_token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
        if not session_token or not request_token:
            abort(400, description='CSRF token ausente.')
        if not secrets.compare_digest(session_token, request_token):
            abort(400, description='CSRF token invalido.')


@app.context_processor
def inject_usuario_logado():
    return {
        'usuario_logado': {
            'id': session.get('usuario_id'),
            'nome': session.get('usuario_nome'),
            'cargo': session.get('usuario_cargo'),
            'role': session.get('role', 'user'),
        },
        'csrf_token': _get_or_create_csrf_token(),
    }


# ---------------------------------------------------------------------------
# Rotas de autenticação
# ---------------------------------------------------------------------------

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        confirmar_senha = request.form.get('confirmar_senha', '')
        cargo = request.form.get('cargo', '').strip()

        if senha != confirmar_senha:
            flash('As senhas não coincidem')
            return redirect('/cadastro')

        usuario_existente = (
            supabase.table('usuarios')
            .select('id')
            .eq('email', email)
            .execute()
        )

        if usuario_existente.data:
            flash('E-mail já cadastrado')
            return redirect('/cadastro')

        senha_hash = generate_password_hash(senha)
        dados = {
            'nome': nome,
            'email': email,
            'senha_hash': senha_hash,
            'cargo': cargo,
            'criado_em': datetime.now(timezone.utc).isoformat(),
        }
        supabase.table('usuarios').insert(dados).execute()
        flash('Cadastro realizado com sucesso!')
        return redirect('/login')

    return render_template('cadastro.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        resposta = (
            supabase.table('usuarios')
            .select('*')
            .eq('email', email)
            .execute()
        )

        usuario = resposta.data[0] if resposta.data else None
        if not usuario or not check_password_hash(usuario['senha_hash'], senha):
            flash('E-mail ou senha incorretos')
            return redirect('/login')

        session['usuario_id'] = usuario['id']
        session['usuario_nome'] = usuario['nome']
        session['usuario_cargo'] = usuario['cargo']
        session['role'] = usuario.get('role', 'user')
        session.permanent = True

        flash(f"Bem-vindo, {usuario['nome']}!")
        return redirect('/')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# ---------------------------------------------------------------------------
# Rotas principais
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def index():
    filtros = _parse_filtros_listagem(request.args)
    query = supabase.table('demandas').select('*,assignee:assignee_id(nome)', count='exact')

    if filtros['filtro_prioridade'] in PRIORIDADES_VALIDAS:
        query = query.eq('prioridade', filtros['filtro_prioridade'])
    if filtros['filtro_solicitante']:
        query = query.eq('solicitante', filtros['filtro_solicitante'])
    if filtros['status_filtro']:
        query = query.in_('status', filtros['status_filtro'])
    if filtros['periodo_inicio']:
        query = query.gte('data_criacao', f"{filtros['periodo_inicio']}T00:00:00")
    if filtros['periodo_fim']:
        query = query.lte('data_criacao', f"{filtros['periodo_fim']}T23:59:59")
    if filtros['assignee_id'].isdigit():
        query = query.eq('assignee_id', int(filtros['assignee_id']))
    if filtros['minhas_demandas']:
        usuario_id = session.get('usuario_id')
        query = query.or_(f'usuario_id.eq.{usuario_id},assignee_id.eq.{usuario_id}')

    coluna_ordenacao = ORDENACOES_LISTAGEM[filtros['sort_by']]
    query = query.order(coluna_ordenacao, desc=filtros['sort_dir'] == 'desc', nullsfirst=False)

    inicio = (filtros['page'] - 1) * filtros['per_page']
    fim = inicio + filtros['per_page'] - 1
    res = query.range(inicio, fim).execute()

    total_registros = res.count or 0
    total_paginas = max((total_registros + filtros['per_page'] - 1) // filtros['per_page'], 1)
    if filtros['page'] > total_paginas:
        filtros['page'] = total_paginas
        inicio = (filtros['page'] - 1) * filtros['per_page']
        fim = inicio + filtros['per_page'] - 1
        res = query.range(inicio, fim).execute()

    dados = preparar_demandas(res.data)

    base_qs = request.args.to_dict(flat=False)
    base_qs.pop('page', None)

    def _build_page_url(page_number):
        params = {k: list(v) for k, v in base_qs.items()}
        params['page'] = [str(page_number)]
        return '/?' + urlencode(params, doseq=True)

    prev_page_url = _build_page_url(filtros['page'] - 1) if filtros['page'] > 1 else None
    next_page_url = _build_page_url(filtros['page'] + 1) if filtros['page'] < total_paginas else None

    solicitantes_res = supabase.table('demandas').select('solicitante').execute()
    solicitantes = listar_solicitantes(solicitantes_res.data)
    usuarios = listar_usuarios()

    return render_template(
        'index.html',
        demandas=dados,
        filtro=filtros['filtro_prioridade'],
        solicitante=filtros['filtro_solicitante'],
        status_filtro=filtros['status_filtro'],
        data_inicio=filtros['periodo_inicio'],
        data_fim=filtros['periodo_fim'],
        assignee_id_filtro=filtros['assignee_id'],
        minhas_demandas=filtros['minhas_demandas'],
        sort_by=filtros['sort_by'],
        sort_dir=filtros['sort_dir'],
        page=filtros['page'],
        per_page=filtros['per_page'],
        total_paginas=total_paginas,
        total_registros=total_registros,
        prev_page_url=prev_page_url,
        next_page_url=next_page_url,
        solicitantes=solicitantes,
        usuarios=usuarios,
        status_validos=STATUS_DEMANDA_VALIDOS,
        prioridades=PRIORIDADES_VALIDAS,
        dias_parada_limite=DIAS_PARADA,
    )


@app.route('/demandas/lote/status', methods=['POST'])
@login_required
def atualizar_status_lote():
    ids = request.form.getlist('demanda_ids')
    novo_status = request.form.get('novo_status', '').strip()
    redirect_to = request.form.get('redirect_to', '').strip()

    def _redirect_listagem():
        if redirect_to.startswith('/'):
            return redirect(redirect_to)
        return redirect('/')

    if not ids:
        flash('Selecione ao menos uma demanda para atualizar em lote.')
        return _redirect_listagem()

    if novo_status not in STATUS_DEMANDA_VALIDOS:
        flash('Status inválido para atualização em lote.')
        return _redirect_listagem()

    ids_validos = []
    for item in ids:
        if str(item).isdigit():
            ids_validos.append(int(item))

    if not ids_validos:
        flash('Nenhuma demanda válida foi selecionada.')
        return _redirect_listagem()

    resposta = (
        supabase
        .table('demandas')
        .select('id,status,usuario_id,assignee_id')
        .in_('id', ids_validos)
        .execute()
    )
    demandas = resposta.data or []

    atualizadas = 0
    sem_permissao = 0
    transicao_invalida = 0
    agora_iso = datetime.now(timezone.utc).isoformat()

    for demanda in demandas:
        if not usuario_pode_alterar_status(demanda):
            sem_permissao += 1
            continue

        status_atual = (demanda.get('status') or 'Aberta').strip()
        if not transicao_status_valida(status_atual, novo_status):
            transicao_invalida += 1
            continue

        dados_update = {
            'status': novo_status,
            'updated_at': agora_iso,
        }
        if status_atual != novo_status:
            dados_update['status_updated_at'] = agora_iso
            if novo_status == 'Finalizada':
                dados_update['resolved_at'] = agora_iso
            elif status_atual == 'Finalizada' and novo_status == 'Aberta':
                dados_update['resolved_at'] = None

        supabase.table('demandas').update(dados_update).eq('id', demanda['id']).execute()
        registrar_evento_demanda(
            demanda_id=demanda['id'],
            tipo='status_alterado_lote',
            before_data={'status': status_atual},
            after_data={'status': novo_status},
        )
        atualizadas += 1

    if atualizadas:
        flash(f'{atualizadas} demanda(s) atualizada(s) em lote.')
    if sem_permissao:
        flash(f'{sem_permissao} demanda(s) ignorada(s) por falta de permissão.')
    if transicao_invalida:
        flash(f'{transicao_invalida} demanda(s) ignorada(s) por transição de status inválida.')

    return _redirect_listagem()


@app.route('/dashboard')
@login_required
def dashboard_redirect():
    return redirect('/')


@app.route('/gerencial/dashboard')
@login_required
@manager_required
def gerencial_dashboard():
    status_ordem = ['Aberta', 'Em andamento', 'Parada', 'Finalizada']

    demandas_res = supabase.table('demandas').select('*,assignee:assignee_id(nome)').execute()
    demandas = preparar_demandas(demandas_res.data or [])

    total_por_status = {status: 0 for status in status_ordem}
    for demanda in demandas:
        status = demanda.get('status') or 'Aberta'
        if status in total_por_status:
            total_por_status[status] += 1

    atraso = []
    for demanda in demandas:
        if not demanda.get('esta_atrasada'):
            continue

        atraso.append({
            'id': demanda.get('id'),
            'titulo': demanda.get('titulo'),
            'responsavel': demanda.get('assignee_nome', 'Nao atribuido'),
            'prioridade': demanda.get('prioridade', 'Média'),
            'dias_atraso': demanda.get('dias_atraso', 0),
            'due_date_fmt': demanda.get('due_date_fmt', '-'),
        })

    atraso = sorted(atraso, key=lambda item: item['dias_atraso'], reverse=True)

    por_usuario_res = (
        supabase.table('demandas')
        .select('assignee_id,status,usuarios:assignee_id(nome)')
        .execute()
    )
    por_usuario_raw = por_usuario_res.data or []
    por_usuario = {}
    for item in por_usuario_raw:
        usuario_id = item.get('assignee_id')
        if not usuario_id:
            continue

        usuario = item.get('usuarios') or {}
        nome = usuario.get('nome') if isinstance(usuario, dict) else None
        if not nome:
            nome = f'Usuário #{usuario_id}'

        if usuario_id not in por_usuario:
            por_usuario[usuario_id] = {
                'nome': nome,
                'total': 0,
                'em_andamento': 0,
                'finalizadas': 0,
            }

        por_usuario[usuario_id]['total'] += 1
        status = item.get('status')
        if status == 'Em andamento':
            por_usuario[usuario_id]['em_andamento'] += 1
        if status == 'Finalizada':
            por_usuario[usuario_id]['finalizadas'] += 1

    demandas_por_usuario = sorted(
        por_usuario.values(),
        key=lambda item: (-item['total'], item['nome'].lower())
    )

    ultimas_res = (
        supabase.table('demandas')
        .select('id,titulo,status,prioridade,solicitante,assignee_id,assignee:assignee_id(nome),updated_at,data_criacao')
        .order('updated_at', desc=True)
        .limit(10)
        .execute()
    )
    ultimas_atualizadas = preparar_demandas(ultimas_res.data or [])

    total_demandas = len(demandas)

    return render_template(
        'gerencial/dashboard.html',
        total_demandas=total_demandas,
        total_por_status=total_por_status,
        demandas_atraso=atraso,
        demandas_por_usuario=demandas_por_usuario,
        ultimas_atualizadas=ultimas_atualizadas,
        chart_labels=status_ordem,
        chart_values=[total_por_status[status] for status in status_ordem],
    )


@app.route('/nova_demanda', methods=['GET', 'POST'])
@login_required
def nova_demanda():
    usuarios = listar_usuarios()
    usuarios_ids = {usuario.get('id') for usuario in usuarios}
    if request.method == 'POST':
        prioridade = request.form.get('prioridade', 'Média')
        if prioridade not in PRIORIDADES_VALIDAS:
            flash('Prioridade inválida.')
            return redirect('/nova_demanda')

        assignee_id = session['usuario_id']
        if session.get('role') == 'manager':
            assignee_form = request.form.get('assignee_id', '').strip()
            if assignee_form:
                try:
                    assignee_id = int(assignee_form)
                except ValueError:
                    flash('Responsavel executor invalido.')
                    return redirect('/nova_demanda')
                if assignee_id not in usuarios_ids:
                    flash('Responsavel executor nao encontrado.')
                    return redirect('/nova_demanda')

        dados = {
            'titulo':      request.form['titulo'],
            'descricao':   request.form['descricao'],
            'solicitante': session['usuario_nome'],
            'prioridade':  prioridade,
            'status':      'Aberta',
            'usuario_id':  session['usuario_id'],
            'assignee_id': assignee_id,
            'status_updated_at': datetime.now(timezone.utc).isoformat(),
            'due_date': calcular_due_date(prioridade).isoformat(),
            'resolved_at': None,
        }
        resposta = supabase.table('demandas').insert(dados).execute()
        demanda_criada = resposta.data[0] if resposta.data else None
        if demanda_criada:
            registrar_evento_demanda(
                demanda_id=demanda_criada.get('id'),
                tipo='criada',
                before_data={},
                after_data={
                    'status': demanda_criada.get('status', 'Aberta'),
                    'prioridade': demanda_criada.get('prioridade'),
                    'assignee_id': demanda_criada.get('assignee_id'),
                },
            )
        flash('Demanda criada com sucesso!')
        return redirect('/')
    return render_template(
        'nova_demanda.html',
        prioridades=PRIORIDADES_VALIDAS,
        usuarios=usuarios,
    )


@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    demanda_atual = buscar_demanda(id)
    if not demanda_atual:
        flash('Demanda não encontrada.')
        return redirect('/')

    pode_gerenciar = usuario_pode_gerenciar(demanda_atual)
    pode_alterar_status = usuario_pode_alterar_status(demanda_atual)
    usuarios = listar_usuarios() if session.get('role') == 'manager' else []
    usuarios_ids = {usuario.get('id') for usuario in usuarios}

    if request.method == 'POST':
        if not pode_alterar_status:
            flash('Você não tem permissão para alterar o status desta demanda.')
            return redirect(f'/editar/{id}')

        status_atual = (demanda_atual.get('status') or 'Aberta').strip()
        novo_status = request.form.get('status', demanda_atual.get('status') or 'Aberta').strip()

        if novo_status not in STATUS_DEMANDA_VALIDOS:
            flash('Status inválido.')
            return redirect(f'/editar/{id}')

        if not transicao_status_valida(status_atual, novo_status):
            flash(f'Transição de status inválida: {status_atual} -> {novo_status}.')
            return redirect(f'/editar/{id}')

        agora = datetime.now(timezone.utc)
        dados = {
            'status':      novo_status,
            'updated_at':  agora.isoformat(),
        }
        eventos_para_registrar = []

        if pode_gerenciar:
            nova_prioridade = request.form.get('prioridade', demanda_atual['prioridade'])

            if ORDEM_PRIORIDADE.get(nova_prioridade, 2) < ORDEM_PRIORIDADE.get(demanda_atual['prioridade'], 2):
                flash('Não é permitido aumentar a prioridade.')
                return redirect(f'/editar/{id}')

            if nova_prioridade not in PRIORIDADES_VALIDAS:
                flash('Prioridade inválida.')
                return redirect(f'/editar/{id}')

            dados.update({
                'titulo':      request.form['titulo'],
                'descricao':   request.form['descricao'],
                'solicitante': request.form['solicitante'],
                'prioridade':  nova_prioridade,
            })

            if nova_prioridade != demanda_atual.get('prioridade'):
                data_base = _parse_iso_datetime(demanda_atual.get('data_criacao')) or agora
                dados['due_date'] = calcular_due_date(nova_prioridade, data_base).isoformat()

            assignee_form = request.form.get('assignee_id', '').strip()
            if assignee_form:
                try:
                    dados['assignee_id'] = int(assignee_form)
                except ValueError:
                    flash('Responsavel executor invalido.')
                    return redirect(f'/editar/{id}')
                if dados['assignee_id'] not in usuarios_ids:
                    flash('Responsavel executor nao encontrado.')
                    return redirect(f'/editar/{id}')

        status_original = demanda_atual.get('status') or 'Aberta'
        prioridade_original = demanda_atual.get('prioridade')
        assignee_original = demanda_atual.get('assignee_id')

        status_final = dados.get('status', status_original)
        prioridade_final = dados.get('prioridade', prioridade_original)
        assignee_final = dados.get('assignee_id', assignee_original)

        if status_final != status_original:
            dados['status_updated_at'] = agora.isoformat()
            if status_final == 'Finalizada':
                dados['resolved_at'] = agora.isoformat()
            elif status_original == 'Finalizada' and status_final == 'Aberta':
                dados['resolved_at'] = None

        if status_final != status_original:
            tipo_evento_status = 'reaberta' if status_original == 'Finalizada' and status_final == 'Aberta' else 'status_alterado'
            eventos_para_registrar.append({
                'tipo': tipo_evento_status,
                'before_data': {'status': status_original},
                'after_data': {'status': status_final},
            })

        if prioridade_final != prioridade_original:
            eventos_para_registrar.append({
                'tipo': 'prioridade_alterada',
                'before_data': {'prioridade': prioridade_original},
                'after_data': {'prioridade': prioridade_final},
            })

        if assignee_final != assignee_original:
            eventos_para_registrar.append({
                'tipo': 'assignee_alterado',
                'before_data': {'assignee_id': assignee_original},
                'after_data': {'assignee_id': assignee_final},
            })

        supabase.table('demandas').update(dados).eq('id', id).execute()
        for evento in eventos_para_registrar:
            registrar_evento_demanda(
                demanda_id=id,
                tipo=evento['tipo'],
                before_data=evento['before_data'],
                after_data=evento['after_data'],
            )
        flash('Demanda atualizada!')
        return redirect('/')

    return render_template(
        'editar.html',
        demanda=demanda_atual,
        pode_gerenciar=pode_gerenciar,
        pode_alterar_status=pode_alterar_status,
        prioridades=PRIORIDADES_VALIDAS,
        usuarios=usuarios,
        status_validos=STATUS_DEMANDA_VALIDOS,
        status_transitions=STATUS_TRANSITIONS,
    )


@app.route('/deletar/<int:id>', methods=['POST'])
@login_required
def deletar(id):
    demanda = buscar_demanda(id)
    if not demanda:
        flash('Demanda não encontrada.')
        return redirect('/')

    if not usuario_pode_gerenciar(demanda):
        flash('Você não pode excluir demanda de outro usuário.')
        return redirect('/')

    supabase.table('demandas').delete().eq('id', id).execute()
    flash('Demanda deletada!')
    return redirect('/')


@app.route('/buscar')
@login_required
def buscar():
    termo = request.args.get('q', '')
    res = supabase.table('demandas').select('*,assignee:assignee_id(nome)').ilike('titulo', f'%{termo}%').execute()
    todos = supabase.table('demandas').select('solicitante').execute()

    dados = preparar_demandas(res.data)

    return render_template(
        'index.html',
        demandas=dados,
        filtro='Todas',
        solicitante='',
        status_filtro=[],
        data_inicio='',
        data_fim='',
        assignee_id_filtro='',
        minhas_demandas=False,
        sort_by='updated_at',
        sort_dir='desc',
        page=1,
        per_page=len(dados) if dados else ITENS_POR_PAGINA_PADRAO,
        total_paginas=1,
        total_registros=len(dados),
        solicitantes=listar_solicitantes(todos.data),
        usuarios=listar_usuarios(),
        status_validos=STATUS_DEMANDA_VALIDOS,
        prioridades=PRIORIDADES_VALIDAS,
        dias_parada_limite=DIAS_PARADA
    )


# ---------------------------------------------------------------------------
# Relatórios — visualização + exportação
# ---------------------------------------------------------------------------

def _coletar_dados_relatorio(filtro_prioridade, filtro_solicitante):
    """Busca e filtra demandas para o relatório. Retorna um dict com tudo."""
    todas_res = supabase.table('demandas').select('*,assignee:assignee_id(nome)').execute()
    todas_demandas = todas_res.data or []
    solicitantes = listar_solicitantes(todas_demandas)
    resumo_solicitantes = gerar_relatorio_solicitantes(todas_demandas)

    demandas_filtradas = todas_demandas
    if filtro_prioridade in PRIORIDADES_VALIDAS:
        demandas_filtradas = [
            d for d in demandas_filtradas if d.get('prioridade') == filtro_prioridade
        ]
    if filtro_solicitante:
        demandas_filtradas = [
            d for d in demandas_filtradas if d.get('solicitante') == filtro_solicitante
        ]

    todos_preparados = preparar_demandas(todas_demandas)
    total_paradas = sum(1 for d in todos_preparados if d.get('dias_parada', 0) >= DIAS_PARADA)
    criticas = filtrar_criticas(todos_preparados)
    total_atrasadas = sum(1 for d in todos_preparados if d.get('esta_atrasada'))

    return {
        'todas_demandas': todas_demandas,
        'demandas_filtradas': preparar_demandas(demandas_filtradas),
        'resumo_solicitantes': resumo_solicitantes,
        'solicitantes': solicitantes,
        'total_demandas': len(todas_demandas),
        'total_solicitantes': len(solicitantes),
        'total_paradas': total_paradas,
        'criticas': criticas,
        'total_criticas': len(criticas),
        'total_atrasadas': total_atrasadas,
    }


@app.route('/relatorios')
@login_required
def relatorios():
    filtro_prioridade  = request.args.get('prioridade', 'Todas')
    filtro_solicitante = request.args.get('solicitante', '').strip()

    dados = _coletar_dados_relatorio(filtro_prioridade, filtro_solicitante)

    return render_template(
        'relatorios.html',
        resumo_solicitantes=dados['resumo_solicitantes'],
        demandas=dados['demandas_filtradas'],
        solicitantes=dados['solicitantes'],
        prioridades=PRIORIDADES_VALIDAS,
        filtro_prioridade=filtro_prioridade,
        filtro_solicitante=filtro_solicitante,
        total_demandas=dados['total_demandas'],
        total_solicitantes=dados['total_solicitantes'],
        total_paradas=dados['total_paradas'],
        total_atrasadas=dados['total_atrasadas'],
        dias_parada_limite=DIAS_PARADA,
        criticas=dados['criticas'],
        total_criticas=dados['total_criticas'],
    )


# ── Exportar CSV ─────────────────────────────────────────────────────────────

@app.route('/relatorios/exportar/csv')
@login_required
def exportar_csv():
    filtro_prioridade  = request.args.get('prioridade', 'Todas')
    filtro_solicitante = request.args.get('solicitante', '').strip()
    dados = _coletar_dados_relatorio(filtro_prioridade, filtro_solicitante)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # --- Seção 1: Resumo geral ---
    writer.writerow(['RELATÓRIO DE DEMANDAS'])
    writer.writerow([f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    writer.writerow([])
    writer.writerow(['RESUMO GERAL'])
    writer.writerow(['Total de demandas', dados['total_demandas']])
    writer.writerow(['Total de solicitantes', dados['total_solicitantes']])
    writer.writerow([f'Demandas paradas (≥ {DIAS_PARADA} dias)', dados['total_paradas']])
    writer.writerow(['Demandas atrasadas (SLA)', dados['total_atrasadas']])
    writer.writerow([])

    # --- Seção 2: Resumo por solicitante ---
    writer.writerow(['RESUMO POR SOLICITANTE'])
    writer.writerow(['Solicitante', 'Total', 'Alta', 'Média', 'Baixa', 'Paradas', 'Atrasadas'])
    for item in dados['resumo_solicitantes']:
        writer.writerow([
            item['solicitante'],
            item['total'],
            item['alta'],
            item['media'],
            item['baixa'],
            item['paradas'],
            item['atrasadas'],
        ])
    writer.writerow([])

    # --- Seção 3: Demandas detalhadas ---
    writer.writerow(['DEMANDAS DETALHADAS'])
    if filtro_prioridade != 'Todas' or filtro_solicitante:
        filtros_ativos = []
        if filtro_prioridade != 'Todas':
            filtros_ativos.append(f'Prioridade: {filtro_prioridade}')
        if filtro_solicitante:
            filtros_ativos.append(f'Solicitante: {filtro_solicitante}')
        writer.writerow([f'Filtros aplicados: {" | ".join(filtros_ativos)}'])

    writer.writerow([
        'ID', 'Título', 'Descrição', 'Solicitante', 'Responsável Executor', 'Prioridade', 'Status',
        'Data de Criação', 'Vencimento SLA', 'Dias Parada', 'Em Atraso', 'Dias Atraso'
    ])
    for d in dados['demandas_filtradas']:
        writer.writerow([
            d.get('id', ''),
            d.get('titulo', ''),
            d.get('descricao', ''),
            d.get('solicitante', ''),
            d.get('assignee_nome', ''),
            d.get('prioridade', ''),
            d.get('status', ''),
            _formatar_data_hora(d.get('data_criacao')),
            d.get('due_date_fmt', '-'),
            d.get('dias_parada', 0),
            'Sim' if d.get('esta_atrasada') else 'Nao',
            d.get('dias_atraso', 0),
        ])

    # Resposta com BOM UTF-8 para o Excel abrir corretamente
    csv_bytes = '\ufeff' + output.getvalue()
    response = make_response(csv_bytes.encode('utf-8'))
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_demandas_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    )
    return response


# ── Exportar Excel (.xlsx) ───────────────────────────────────────────────────

@app.route('/relatorios/exportar/excel')
@login_required
def exportar_excel():
    filtro_prioridade  = request.args.get('prioridade', 'Todas')
    filtro_solicitante = request.args.get('solicitante', '').strip()
    dados = _coletar_dados_relatorio(filtro_prioridade, filtro_solicitante)

    wb = Workbook()

    # Estilos reutilizáveis
    cor_cabecalho   = '2563EB'   # azul
    cor_alta        = 'FEE2E2'   # vermelho claro
    cor_media       = 'FEF9C3'   # amarelo claro
    cor_baixa       = 'DCFCE7'   # verde claro
    cor_parada      = 'FCA5A5'   # vermelho mais forte (paradas)
    cor_resumo_bg   = 'EFF6FF'   # azul muito claro

    fonte_titulo  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    fonte_sub     = Font(name='Arial', bold=True, size=11)
    fonte_cabecalho = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    fonte_normal  = Font(name='Arial', size=10)
    fonte_negrito = Font(name='Arial', bold=True, size=10)

    fill_cabecalho = PatternFill('solid', start_color=cor_cabecalho)
    fill_resumo    = PatternFill('solid', start_color=cor_resumo_bg)

    borda_fina = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ── Aba 1: Resumo ──────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'
    ws_resumo.sheet_view.showGridLines = False

    # Título principal
    ws_resumo.merge_cells('A1:F1')
    ws_resumo['A1'] = 'RELATÓRIO DE DEMANDAS'
    ws_resumo['A1'].font = fonte_titulo
    ws_resumo['A1'].fill = fill_cabecalho
    ws_resumo['A1'].alignment = centro
    ws_resumo.row_dimensions[1].height = 30

    # Data de geração
    ws_resumo.merge_cells('A2:F2')
    ws_resumo['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
    ws_resumo['A2'].font = Font(name='Arial', italic=True, size=9, color='6B7280')
    ws_resumo['A2'].alignment = centro
    ws_resumo.row_dimensions[2].height = 16

    ws_resumo.append([])  # linha vazia (linha 3)

    # Cards de resumo geral (linha 4-5)
    ws_resumo.merge_cells('A4:B5')
    ws_resumo['A4'] = 'Total de Demandas'
    ws_resumo['A4'].font = fonte_negrito
    ws_resumo['A4'].alignment = centro
    ws_resumo['A4'].fill = fill_resumo
    ws_resumo['A4'].border = borda_fina

    ws_resumo.merge_cells('C4:D5')
    ws_resumo['C4'] = 'Total de Solicitantes'
    ws_resumo['C4'].font = fonte_negrito
    ws_resumo['C4'].alignment = centro
    ws_resumo['C4'].fill = fill_resumo
    ws_resumo['C4'].border = borda_fina

    ws_resumo.merge_cells('E4:F5')
    ws_resumo['E4'] = 'Demandas Paradas'
    ws_resumo['E4'].font = fonte_negrito
    ws_resumo['E4'].alignment = centro
    ws_resumo['E4'].fill = fill_resumo
    ws_resumo['E4'].border = borda_fina

    ws_resumo.merge_cells('A6:B7')
    ws_resumo['A6'] = dados['total_demandas']
    ws_resumo['A6'].font = Font(name='Arial', bold=True, size=22, color=cor_cabecalho)
    ws_resumo['A6'].alignment = centro
    ws_resumo['A6'].border = borda_fina

    ws_resumo.merge_cells('C6:D7')
    ws_resumo['C6'] = dados['total_solicitantes']
    ws_resumo['C6'].font = Font(name='Arial', bold=True, size=22, color='059669')
    ws_resumo['C6'].alignment = centro
    ws_resumo['C6'].border = borda_fina

    ws_resumo.merge_cells('E6:F7')
    ws_resumo['E6'] = dados['total_paradas']
    ws_resumo['E6'].font = Font(name='Arial', bold=True, size=22, color='DC2626')
    ws_resumo['E6'].alignment = centro
    ws_resumo['E6'].border = borda_fina

    ws_resumo.row_dimensions[4].height = 20
    ws_resumo.row_dimensions[5].height = 20
    ws_resumo.row_dimensions[6].height = 28
    ws_resumo.row_dimensions[7].height = 28

    ws_resumo.append([])  # linha 8 vazia

    # Tabela de resumo por solicitante
    linha_sub = 9
    ws_resumo.merge_cells(f'A{linha_sub}:F{linha_sub}')
    ws_resumo[f'A{linha_sub}'] = 'RESUMO POR SOLICITANTE'
    ws_resumo[f'A{linha_sub}'].font = fonte_sub
    ws_resumo[f'A{linha_sub}'].alignment = esquerda
    ws_resumo.row_dimensions[linha_sub].height = 20

    linha_cab = linha_sub + 1
    cabecalhos_resumo = ['Solicitante', 'Total', 'Alta', 'Média', 'Baixa', f'Paradas (≥{DIAS_PARADA} dias)']
    for col, titulo in enumerate(cabecalhos_resumo, start=1):
        cel = ws_resumo.cell(row=linha_cab, column=col, value=titulo)
        cel.font = fonte_cabecalho
        cel.fill = fill_cabecalho
        cel.alignment = centro
        cel.border = borda_fina
    ws_resumo.row_dimensions[linha_cab].height = 22

    for item in dados['resumo_solicitantes']:
        linha_cab += 1
        valores = [
            item['solicitante'], item['total'], item['alta'],
            item['media'], item['baixa'], item['paradas']
        ]
        for col, val in enumerate(valores, start=1):
            cel = ws_resumo.cell(row=linha_cab, column=col, value=val)
            cel.font = fonte_normal
            cel.alignment = centro if col > 1 else esquerda
            cel.border = borda_fina
        ws_resumo.row_dimensions[linha_cab].height = 18

    # Larguras das colunas — aba Resumo
    larguras_resumo = [30, 10, 10, 10, 10, 18]
    for i, larg in enumerate(larguras_resumo, start=1):
        ws_resumo.column_dimensions[get_column_letter(i)].width = larg

    # ── Aba 2: Demandas detalhadas ─────────────────────────────────────────
    ws_det = wb.create_sheet('Demandas Detalhadas')
    ws_det.sheet_view.showGridLines = False

    # Título
    ws_det.merge_cells('A1:H1')
    ws_det['A1'] = 'DEMANDAS DETALHADAS'
    ws_det['A1'].font = fonte_titulo
    ws_det['A1'].fill = fill_cabecalho
    ws_det['A1'].alignment = centro
    ws_det.row_dimensions[1].height = 30

    # Filtros ativos
    filtros_txt = []
    if filtro_prioridade != 'Todas':
        filtros_txt.append(f'Prioridade: {filtro_prioridade}')
    if filtro_solicitante:
        filtros_txt.append(f'Solicitante: {filtro_solicitante}')

    ws_det.merge_cells('A2:H2')
    ws_det['A2'] = ('Filtros: ' + ' | '.join(filtros_txt)) if filtros_txt else 'Sem filtros aplicados'
    ws_det['A2'].font = Font(name='Arial', italic=True, size=9, color='6B7280')
    ws_det['A2'].alignment = centro
    ws_det.row_dimensions[2].height = 16

    ws_det.append([])  # linha 3

    # Cabeçalhos
    cabecalhos_det = [
        'ID', 'Título', 'Descrição', 'Solicitante', 'Responsável Executor', 'Prioridade', 'Status',
        'Data de Criação', 'Vencimento SLA', 'Dias Parada', 'Em Atraso', 'Dias Atraso'
    ]
    for col, titulo in enumerate(cabecalhos_det, start=1):
        cel = ws_det.cell(row=4, column=col, value=titulo)
        cel.font = fonte_cabecalho
        cel.fill = fill_cabecalho
        cel.alignment = centro
        cel.border = borda_fina
    ws_det.row_dimensions[4].height = 22

    # Mapa de cores de prioridade
    fill_alta  = PatternFill('solid', start_color=cor_alta)
    fill_media = PatternFill('solid', start_color=cor_media)
    fill_baixa = PatternFill('solid', start_color=cor_baixa)
    fill_parada = PatternFill('solid', start_color=cor_parada)

    for idx, d in enumerate(dados['demandas_filtradas'], start=5):
        prioridade = d.get('prioridade', '')
        dias_parada = d.get('dias_parada', 0)

        # Escolhe fill de linha conforme prioridade (vermelho mais forte se parada há 3+ dias)
        if dias_parada >= DIAS_PARADA:
            fill_linha = fill_parada
        elif prioridade == 'Alta':
            fill_linha = fill_alta
        elif prioridade == 'Média':
            fill_linha = fill_media
        else:
            fill_linha = fill_baixa

        valores = [
            d.get('id', ''),
            d.get('titulo', ''),
            d.get('descricao', ''),
            d.get('solicitante', ''),
            d.get('assignee_nome', ''),
            prioridade,
            d.get('status', ''),
            _formatar_data_hora(d.get('data_criacao')),
            d.get('due_date_fmt', '-'),
            dias_parada,
            'Sim' if d.get('esta_atrasada') else 'Nao',
            d.get('dias_atraso', 0),
        ]
        for col, val in enumerate(valores, start=1):
            cel = ws_det.cell(row=idx, column=col, value=val)
            cel.font = fonte_normal
            cel.fill = fill_linha
            cel.border = borda_fina
            cel.alignment = centro if col in (1, 6, 7, 8, 10, 11, 12) else esquerda
        ws_det.row_dimensions[idx].height = 18

    # Larguras das colunas — aba Demandas
    larguras_det = [8, 28, 32, 18, 20, 12, 14, 18, 18, 12, 12, 12]
    for i, larg in enumerate(larguras_det, start=1):
        ws_det.column_dimensions[get_column_letter(i)].width = larg

    # ── Salvar em buffer ───────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_demandas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    return response


# ── Exportar PDF ─────────────────────────────────────────────────────────────

@app.route('/relatorios/exportar/pdf')
@login_required
def exportar_pdf():
    filtro_prioridade  = request.args.get('prioridade', 'Todas')
    filtro_solicitante = request.args.get('solicitante', '').strip()
    dados = _coletar_dados_relatorio(filtro_prioridade, filtro_solicitante)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    cor_azul = colors.HexColor('#2563EB')
    cor_cinza = colors.HexColor('#6B7280')

    estilo_titulo = ParagraphStyle(
        'Titulo',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.white,
        alignment=TA_CENTER,
        spaceAfter=0,
    )
    estilo_sub = ParagraphStyle(
        'Sub',
        parent=styles['Normal'],
        fontSize=11,
        textColor=cor_azul,
        fontName='Helvetica-Bold',
        spaceBefore=12,
        spaceAfter=4,
    )
    estilo_meta = ParagraphStyle(
        'Meta',
        parent=styles['Normal'],
        fontSize=8,
        textColor=cor_cinza,
        alignment=TA_CENTER,
    )

    elementos = []

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    titulo_tabela = Table(
        [[Paragraph('RELATÓRIO DE DEMANDAS', estilo_titulo)]],
        colWidths=[doc.width],
    )
    titulo_tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), cor_azul),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('ROUNDEDCORNERS', [4]),
    ]))
    elementos.append(titulo_tabela)
    elementos.append(Spacer(1, 0.2 * cm))
    elementos.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")}', estilo_meta
    ))

    # Filtros ativos
    filtros_txt = []
    if filtro_prioridade != 'Todas':
        filtros_txt.append(f'Prioridade: {filtro_prioridade}')
    if filtro_solicitante:
        filtros_txt.append(f'Solicitante: {filtro_solicitante}')
    if filtros_txt:
        elementos.append(Paragraph(
            f'Filtros aplicados: {" | ".join(filtros_txt)}', estilo_meta
        ))

    elementos.append(Spacer(1, 0.4 * cm))

    # ── Cards de resumo ────────────────────────────────────────────────────
    estilo_card_label = ParagraphStyle('CardLabel', parent=styles['Normal'],
                                       fontSize=9, textColor=cor_cinza,
                                       alignment=TA_CENTER)
    estilo_card_valor = ParagraphStyle('CardValor', parent=styles['Normal'],
                                       fontSize=20, fontName='Helvetica-Bold',
                                       alignment=TA_CENTER)

    cards_data = [[
        Paragraph('Total de Demandas', estilo_card_label),
        Paragraph('Total de Solicitantes', estilo_card_label),
        Paragraph('Demandas Paradas', estilo_card_label),
        Paragraph('Atrasadas (SLA)', estilo_card_label),
    ], [
        Paragraph(str(dados['total_demandas']), estilo_card_valor),
        Paragraph(str(dados['total_solicitantes']), estilo_card_valor),
        Paragraph(str(dados['total_paradas']), estilo_card_valor),
        Paragraph(str(dados['total_atrasadas']), estilo_card_valor),
    ]]

    larg_card = doc.width / 4
    tabela_cards = Table(cards_data, colWidths=[larg_card] * 4, rowHeights=[16, 32])
    tabela_cards.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFF6FF')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('LINEAFTER', (0, 0), (2, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TEXTCOLOR', (0, 1), (0, 1), cor_azul),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#DC2626')),
        ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#B91C1C')),
    ]))
    elementos.append(tabela_cards)
    elementos.append(Spacer(1, 0.5 * cm))

    # ── Tabela: Resumo por solicitante ─────────────────────────────────────
    elementos.append(Paragraph('Resumo por Solicitante', estilo_sub))

    cab_resumo = [['Solicitante', 'Total', 'Alta', 'Média', 'Baixa', 'Paradas (≥3 dias)']]
    linhas_resumo = cab_resumo + [
        [item['solicitante'], item['total'], item['alta'],
         item['media'], item['baixa'], item['paradas']]
        for item in dados['resumo_solicitantes']
    ]

    col_widths_resumo = [
        doc.width * 0.35, doc.width * 0.12, doc.width * 0.12,
        doc.width * 0.12, doc.width * 0.12, doc.width * 0.17,
    ]
    tabela_resumo = Table(linhas_resumo, colWidths=col_widths_resumo, repeatRows=1)
    style_resumo = [
        ('BACKGROUND', (0, 0), (-1, 0), cor_azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]
    tabela_resumo.setStyle(TableStyle(style_resumo))
    elementos.append(tabela_resumo)
    elementos.append(Spacer(1, 0.5 * cm))

    # ── Tabela: Demandas detalhadas ────────────────────────────────────────
    elementos.append(Paragraph('Demandas Detalhadas', estilo_sub))

    estilo_cel = ParagraphStyle('Cel', parent=styles['Normal'], fontSize=7,
                                leading=9, wordWrap='LTR')

    cab_det = [[
        'ID', 'Título', 'Solicitante', 'Responsável Executor', 'Prioridade', 'Status',
        'Data Criação', 'Vencimento SLA', 'Dias Parada', 'Em Atraso'
    ]]

    linhas_det = cab_det.copy()
    for d in dados['demandas_filtradas']:
        linhas_det.append([
            str(d.get('id', '')),
            Paragraph(d.get('titulo', ''), estilo_cel),
            d.get('solicitante', ''),
            d.get('assignee_nome', ''),
            d.get('prioridade', ''),
            d.get('status', ''),
            _formatar_data_hora(d.get('data_criacao')).replace(' ', '\n'),
            d.get('due_date_fmt', '-').replace(' ', '\n'),
            str(d.get('dias_parada', 0)),
            'Sim' if d.get('esta_atrasada') else 'Nao',
        ])

    col_widths_det = [
        doc.width * 0.05,
        doc.width * 0.23,
        doc.width * 0.12,
        doc.width * 0.14,
        doc.width * 0.08,
        doc.width * 0.10,
        doc.width * 0.10,
        doc.width * 0.10,
        doc.width * 0.04,
        doc.width * 0.04,
    ]
    tabela_det = Table(linhas_det, colWidths=col_widths_det, repeatRows=1)

    style_det = [
        ('BACKGROUND', (0, 0), (-1, 0), cor_azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (4, 0), (9, -1), 'CENTER'),
        ('ALIGN', (1, 0), (3, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]

    # Colorir linhas por prioridade / parada
    for row_idx, d in enumerate(dados['demandas_filtradas'], start=1):
        prioridade = d.get('prioridade', '')
        dias_parada = d.get('dias_parada', 0)
        if dias_parada >= DIAS_PARADA:
            bg = colors.HexColor('#FCA5A5')
        elif prioridade == 'Alta':
            bg = colors.HexColor('#FEE2E2')
        elif prioridade == 'Média':
            bg = colors.HexColor('#FEF9C3')
        else:
            bg = colors.HexColor('#DCFCE7')
        style_det.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg))

    tabela_det.setStyle(TableStyle(style_det))
    elementos.append(tabela_det)

    # ── Rodapé via onLaterPages ────────────────────────────────────────────
    def rodape(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(cor_cinza)
        canvas.drawCentredString(
            doc.pagesize[0] / 2,
            0.8 * cm,
            f'Página {doc.page} — Relatório gerado automaticamente pelo sistema de demandas'
        )
        canvas.restoreState()

    doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_demandas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    )
    return response


# ── Exportar Críticas (Alta + Paradas) — Excel ───────────────────────────────

@app.route('/relatorios/exportar/criticas')
@login_required
def exportar_criticas():
    """Exporta em Excel apenas as demandas de prioridade Alta com dias_parada >= DIAS_PARADA."""
    todas_res = supabase.table('demandas').select('*,assignee:assignee_id(nome)').execute()
    todas_preparadas = preparar_demandas(todas_res.data or [])
    criticas = filtrar_criticas(todas_preparadas)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Demandas Críticas'
    ws.sheet_view.showGridLines = False

    cor_vermelho   = 'DC2626'
    cor_vermelho_bg = 'FEF2F2'
    cor_laranja_bg  = 'FCA5A5'

    fonte_titulo   = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    fonte_cab      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    fonte_normal   = Font(name='Arial', size=10)
    fonte_negrito  = Font(name='Arial', bold=True, size=10)

    fill_header  = PatternFill('solid', start_color=cor_vermelho)
    fill_linha   = PatternFill('solid', start_color=cor_vermelho_bg)
    fill_urgente = PatternFill('solid', start_color=cor_laranja_bg)

    borda = Border(
        left=Side(style='thin', color='FECACA'),
        right=Side(style='thin', color='FECACA'),
        top=Side(style='thin', color='FECACA'),
        bottom=Side(style='thin', color='FECACA'),
    )
    centro   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = '⚠ DEMANDAS CRÍTICAS — Alta Prioridade + Paradas'
    ws['A1'].font = fonte_titulo
    ws['A1'].fill = fill_header
    ws['A1'].alignment = centro
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells('A2:H2')
    ws['A2'] = (
        f'Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")}  |  '
        f'Critério: Prioridade Alta + Paradas há ≥ {DIAS_PARADA} dias  |  '
        f'Total: {len(criticas)} demanda(s)'
    )
    ws['A2'].font = Font(name='Arial', italic=True, size=9, color='7F1D1D')
    ws['A2'].alignment = centro
    ws.row_dimensions[2].height = 16

    ws.append([])  # linha 3 vazia

    # Cabeçalhos
    cabecalhos = ['ID', 'Título', 'Descrição', 'Solicitante', 'Responsável Executor', 'Prioridade',
                  'Data de Criação', 'Dias Parada']
    for col, titulo in enumerate(cabecalhos, start=1):
        cel = ws.cell(row=4, column=col, value=titulo)
        cel.font = fonte_cab
        cel.fill = fill_header
        cel.alignment = centro
        cel.border = borda
    ws.row_dimensions[4].height = 22

    if not criticas:
        ws.merge_cells('A5:H5')
        ws['A5'] = 'Nenhuma demanda crítica encontrada no momento.'
        ws['A5'].font = Font(name='Arial', italic=True, size=10, color='6B7280')
        ws['A5'].alignment = centro
    else:
        for idx, d in enumerate(criticas, start=5):
            data_str = d.get('data_criacao', '')
            try:
                dt = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
                data_fmt = dt.strftime('%d/%m/%Y %H:%M')
            except Exception:
                data_fmt = data_str

            dias = d.get('dias_parada', 0)
            # Mais de 7 dias: destaque laranja mais forte
            fill_row = fill_urgente if dias >= 7 else fill_linha

            valores = [
                d.get('id', ''), d.get('titulo', ''), d.get('descricao', ''),
                d.get('solicitante', ''), d.get('assignee_nome', ''), d.get('prioridade', ''), data_fmt, dias,
            ]
            for col, val in enumerate(valores, start=1):
                cel = ws.cell(row=idx, column=col, value=val)
                cel.font = fonte_negrito if col == 7 else fonte_normal
                cel.fill = fill_row
                cel.border = borda
                cel.alignment = centro if col in (1, 6, 7, 8) else esquerda
            ws.row_dimensions[idx].height = 18

    larguras = [8, 30, 40, 20, 22, 12, 18, 12]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = (
        f'attachment; filename="criticas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    return response


# ---------------------------------------------------------------------------
# Detalhes e comentários
# ---------------------------------------------------------------------------

@app.route('/detalhes/<int:id>')
@login_required
def detalhes(id):
    demanda = supabase.table('demandas').select('*,assignee:assignee_id(nome)').eq('id', id).single().execute()
    eventos = listar_eventos_demanda(id)
    comentarios = (
        supabase
        .table('comentarios')
        .select('*,usuarios:autor_id(nome)')
        .eq('demanda_id', id)
        .order('data')
        .execute()
    )
    return render_template(
        'detalhes.html',
        demanda=preparar_demandas([demanda.data])[0],
        eventos=eventos,
        comentarios=comentarios.data,
        pode_gerenciar=usuario_pode_gerenciar(demanda.data),
    )


@app.route('/adicionar_comentario/<int:demanda_id>', methods=['POST'])
@login_required
def adicionar_comentario(demanda_id):
    dados = {
        'demanda_id': demanda_id,
        'comentario': request.form['comentario'],
        'autor':      session.get('usuario_nome'),
        'autor_id':   session.get('usuario_id'),
    }
    supabase.table('comentarios').insert(dados).execute()
    return redirect(f'/detalhes/{demanda_id}')


if __name__ == '__main__':
    app.run(
        debug=_env_to_bool(os.environ.get('FLASK_DEBUG')),
        host='0.0.0.0'
    )
